"""
Builds MachineMonitor.xlsx — the base workbook that ships with all sheets,
tables, sample data, formatting, and named ranges pre-configured.

CSV schema (8 columns):
    machine_id, tag, timestamp, metric1, metric2, metric3,
    running_hours (cumulative meter), downtime_hours (per-row delta)
"""

from datetime import datetime, timedelta
from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation

OUT_XLSX = Path(__file__).parent / "MachineMonitor.xlsx"

# ---------- Palette ----------
INK        = "0F1A20"
PANEL      = "13232B"
PANEL_2    = "1B303A"
DIVIDER    = "24404C"
STEEL      = "6E8A96"
FROST      = "E6EEF1"
PAPER      = "F4F7F8"
AMBER      = "E8A33D"
AMBER_2    = "C87A1F"
LIME       = "9BC53D"
RED        = "D64545"
CHART_A    = "3FA7B8"
CHART_B    = "E8A33D"
CHART_C    = "9BC53D"


def make_styles(wb: Workbook) -> None:
    styles = {
        "hdr_title": NamedStyle(
            name="hdr_title",
            font=Font(name="Consolas", size=22, bold=True, color=FROST),
            fill=PatternFill("solid", fgColor=PANEL),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "hdr_sub": NamedStyle(
            name="hdr_sub",
            font=Font(name="Consolas", size=10, color=STEEL),
            fill=PatternFill("solid", fgColor=PANEL),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "section": NamedStyle(
            name="section",
            font=Font(name="Consolas", size=11, bold=True, color=AMBER),
            fill=PatternFill("solid", fgColor=PANEL_2),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "label": NamedStyle(
            name="label",
            font=Font(name="Consolas", size=10, color=STEEL),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "value": NamedStyle(
            name="value",
            font=Font(name="Consolas", size=11, bold=True, color=INK),
            fill=PatternFill("solid", fgColor=FROST),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=Border(
                left=Side(style="thin", color=DIVIDER),
                right=Side(style="thin", color=DIVIDER),
                top=Side(style="thin", color=DIVIDER),
                bottom=Side(style="thin", color=DIVIDER),
            ),
        ),
        "input": NamedStyle(
            name="input",
            font=Font(name="Consolas", size=10, color=INK),
            fill=PatternFill("solid", fgColor="FFFDF3"),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=Border(
                left=Side(style="thin", color=AMBER_2),
                right=Side(style="thin", color=AMBER_2),
                top=Side(style="thin", color=AMBER_2),
                bottom=Side(style="thin", color=AMBER_2),
            ),
        ),
        "kpi_label": NamedStyle(
            name="kpi_label",
            font=Font(name="Consolas", size=9, color=STEEL),
            fill=PatternFill("solid", fgColor=PANEL_2),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "kpi_val": NamedStyle(
            name="kpi_val",
            font=Font(name="Consolas", size=18, bold=True, color=AMBER),
            fill=PatternFill("solid", fgColor=PANEL_2),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            number_format="0.00",
        ),
        "kpi_val_pct": NamedStyle(
            name="kpi_val_pct",
            font=Font(name="Consolas", size=18, bold=True, color=LIME),
            fill=PatternFill("solid", fgColor=PANEL_2),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            number_format="0.0",
        ),
        "btn": NamedStyle(
            name="btn",
            font=Font(name="Consolas", size=11, bold=True, color=INK),
            fill=PatternFill("solid", fgColor=AMBER),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=Border(
                left=Side(style="medium", color=AMBER_2),
                right=Side(style="medium", color=AMBER_2),
                top=Side(style="medium", color=AMBER_2),
                bottom=Side(style="medium", color=AMBER_2),
            ),
        ),
        "th": NamedStyle(
            name="th",
            font=Font(name="Consolas", size=10, bold=True, color=FROST),
            fill=PatternFill("solid", fgColor=PANEL_2),
            alignment=Alignment(horizontal="center", vertical="center"),
        ),
        "th_left": NamedStyle(
            name="th_left",
            font=Font(name="Consolas", size=10, bold=True, color=FROST),
            fill=PatternFill("solid", fgColor=PANEL_2),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "pick_id": NamedStyle(
            name="pick_id",
            font=Font(name="Consolas", size=10, bold=True, color=INK),
            fill=PatternFill("solid", fgColor=FROST),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=Border(
                left=Side(style="thin", color=DIVIDER),
                right=Side(style="thin", color=DIVIDER),
                top=Side(style="thin", color=DIVIDER),
                bottom=Side(style="thin", color=DIVIDER),
            ),
        ),
        "pick_show": NamedStyle(
            name="pick_show",
            font=Font(name="Consolas", size=10, bold=True, color=INK),
            fill=PatternFill("solid", fgColor="E9F3E4"),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=Border(
                left=Side(style="thin", color=DIVIDER),
                right=Side(style="thin", color=DIVIDER),
                top=Side(style="thin", color=DIVIDER),
                bottom=Side(style="thin", color=DIVIDER),
            ),
        ),
        "util_val": NamedStyle(
            name="util_val",
            font=Font(name="Consolas", size=10, color=INK),
            fill=PatternFill("solid", fgColor="FFFDF3"),
            alignment=Alignment(horizontal="right", vertical="center", indent=1),
            border=Border(
                left=Side(style="thin", color=DIVIDER),
                right=Side(style="thin", color=DIVIDER),
                top=Side(style="thin", color=DIVIDER),
                bottom=Side(style="thin", color=DIVIDER),
            ),
            number_format="0.00",
        ),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def build_config(wb: Workbook) -> None:
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMBER

    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.merge_cells("B1:H1")
    ws["B1"] = "MACHINE  CONFIGURATION"
    ws["B1"].style = "hdr_title"
    ws.merge_cells("B2:H2")
    ws["B2"] = "Register up to 16 machines. Assign a tag to group them for side-by-side comparison."
    ws["B2"].style = "hdr_sub"

    headers = ["Machine ID", "Machine Name", "Tag",
               "Metric 1 Name", "Metric 2 Name", "Metric 3 Name", "Active"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 26

    tags = ["Line-A", "Line-B", "Line-C", "Line-D"]
    metric_sets = [
        ("Temperature (°C)", "Vibration (mm/s)", "Output (units/h)"),
        ("Pressure (bar)",   "Flow (L/min)",     "RPM"),
        ("Torque (Nm)",      "Current (A)",      "Voltage (V)"),
        ("Speed (m/s)",      "Load (%)",         "Cycles/hr"),
    ]
    for row_i in range(16):
        r = 5 + row_i
        tag_idx = row_i // 4
        ws.cell(row=r, column=2, value=f"M{row_i+1:02d}")
        ws.cell(row=r, column=3, value=f"Machine {row_i+1:02d}")
        ws.cell(row=r, column=4, value=tags[tag_idx])
        ws.cell(row=r, column=5, value=metric_sets[tag_idx][0])
        ws.cell(row=r, column=6, value=metric_sets[tag_idx][1])
        ws.cell(row=r, column=7, value=metric_sets[tag_idx][2])
        ws.cell(row=r, column=8, value="Yes")
        for col in range(2, 9):
            ws.cell(row=r, column=col).style = "value"

    widths = {"A": 2, "B": 12, "C": 18, "D": 12, "E": 22, "F": 22, "G": 22, "H": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    tbl = Table(displayName="tblMachines", ref="B4:H20")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("H5:H20")


def build_data(wb: Workbook) -> None:
    ws = wb.create_sheet("Data")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = CHART_A

    ws.row_dimensions[1].height = 46
    ws.merge_cells("B1:I1")
    ws["B1"] = "RAW  DATA  STORE"
    ws["B1"].style = "hdr_title"
    ws.merge_cells("B2:I2")
    ws["B2"] = "CSV columns: machine_id, tag, timestamp, metric1, metric2, metric3, running_hours(cum), downtime_hours(delta)."
    ws["B2"].style = "hdr_sub"

    headers = ["Machine ID", "Tag", "Timestamp",
               "Metric 1", "Metric 2", "Metric 3",
               "Running Hrs", "Downtime Hrs"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 26

    tags = ["Line-A"] * 4 + ["Line-B"] * 4 + ["Line-C"] * 4 + ["Line-D"] * 4
    rng = random.Random(7)
    start = datetime(2026, 1, 1)

    # Cumulative running-hours meter per machine, starts at 1000 h
    run_meter = [1000.0 + m * 137.0 for m in range(16)]

    row = 5
    for day in range(30):
        for m_idx in range(16):
            mid = f"M{m_idx+1:02d}"
            tag = tags[m_idx]
            ts = start + timedelta(days=day, hours=8)
            base = 50 + m_idx * 2
            m1 = round(base + rng.uniform(-5, 5), 2)
            m2 = round(10 + rng.uniform(0, 8), 2)
            m3 = round(100 + m_idx * 5 + rng.uniform(-15, 15), 2)

            # Downtime this reporting period (delta hours): 0..3 h
            dt = round(rng.uniform(0, 3), 2)
            # Advance meter: ran for (24 − downtime) hours since last reading
            run_meter[m_idx] += max(0.0, 24.0 - dt)

            ws.cell(row=row, column=2, value=mid)
            ws.cell(row=row, column=3, value=tag)
            ws.cell(row=row, column=4, value=ts).number_format = "yyyy-mm-dd hh:mm"
            ws.cell(row=row, column=5, value=m1)
            ws.cell(row=row, column=6, value=m2)
            ws.cell(row=row, column=7, value=m3)
            ws.cell(row=row, column=8, value=round(run_meter[m_idx], 2))
            ws.cell(row=row, column=9, value=dt)
            row += 1

    end_row = row - 1
    tbl = Table(displayName="tblData", ref=f"B4:I{end_row}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tbl)

    widths = {"A": 2, "B": 12, "C": 12, "D": 20,
              "E": 12, "F": 12, "G": 12, "H": 14, "I": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = INK

    for col in range(1, 40):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.column_dimensions["A"].width = 2

    # ---------- Title ----------
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 22
    ws.merge_cells("B2:P2")
    ws["B2"] = "MACHINE  MONITORING  ///  CONTROL  ROOM"
    ws["B2"].style = "hdr_title"
    ws.merge_cells("B3:P3")
    ws["B3"] = "CSV → filter → pick machines → compare metrics + utilization → export report."
    ws["B3"].style = "hdr_sub"

    # ---------- Filters ----------
    ws.row_dimensions[5].height = 22
    ws.merge_cells("B5:P5")
    ws["B5"] = "  FILTERS"
    ws["B5"].style = "section"
    ws.row_dimensions[6].height = 8
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 24

    ws["B7"] = "Start Date"; ws["B7"].style = "label"
    ws["B8"] = "End Date";   ws["B8"].style = "label"
    ws.merge_cells("C7:D7"); ws.merge_cells("C8:D8")
    ws["C7"] = datetime(2026, 1, 1);  ws["C7"].number_format = "yyyy-mm-dd"
    ws["C8"] = datetime(2026, 1, 30); ws["C8"].number_format = "yyyy-mm-dd"
    ws["C7"].style = "value"; ws["C8"].style = "value"

    ws["F7"] = "Tag Group"; ws["F7"].style = "label"
    ws.merge_cells("G7:H7"); ws["G7"] = "Line-A"; ws["G7"].style = "value"
    dv_tag = DataValidation(type="list", formula1="=Config!$D$5:$D$20", allow_blank=False)
    ws.add_data_validation(dv_tag)
    dv_tag.add("G7:H7")

    ws["F8"] = "Metric Focus"; ws["F8"].style = "label"
    ws.merge_cells("G8:H8"); ws["G8"] = "Metric 1"; ws["G8"].style = "value"
    dv_metric = DataValidation(type="list", formula1='"Metric 1,Metric 2,Metric 3"', allow_blank=False)
    ws.add_data_validation(dv_metric)
    dv_metric.add("G8:H8")

    ws.merge_cells("J7:L7"); ws["J7"] = "▸  IMPORT CSV";            ws["J7"].style = "btn"
    ws.merge_cells("J8:L8"); ws["J8"] = "▸  REFRESH DASHBOARD";     ws["J8"].style = "btn"
    ws.merge_cells("N7:P7"); ws["N7"] = "▸  GENERATE REPORT";       ws["N7"].style = "btn"
    ws.merge_cells("N8:P8"); ws["N8"] = "▸  CLEAR DATA";            ws["N8"].style = "btn"

    # ---------- KPI row 1 (metrics) ----------
    ws.row_dimensions[10].height = 22
    ws.merge_cells("B10:P10")
    ws["B10"] = "  KEY  METRICS  (filtered)"
    ws["B10"].style = "section"
    ws.row_dimensions[11].height = 18
    ws.row_dimensions[12].height = 34

    kpi_row1 = [
        ("MACHINES",
         "=SUMPRODUCT((Config!D5:D20=G7)*(Config!H5:H20=\"Yes\"))", "0"),
        ("DATA POINTS",
         "=SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1))",
         "#,##0"),
        ("AVG METRIC 1",
         "=IFERROR(SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1)*Data!E5:E20000)"
         "/MAX(1,SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1))),0)", "0.00"),
        ("AVG METRIC 2",
         "=IFERROR(SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1)*Data!F5:F20000)"
         "/MAX(1,SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1))),0)", "0.00"),
        ("AVG METRIC 3",
         "=IFERROR(SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1)*Data!G5:G20000)"
         "/MAX(1,SUMPRODUCT((Data!C5:C20000=G7)*(Data!D5:D20000>=C7)*(Data!D5:D20000<=C8+1))),0)", "0.00"),
    ]
    for (label, formula, fmt), col in zip(kpi_row1, ["B", "E", "H", "K", "N"]):
        end_col = get_column_letter(ord(col) - ord("A") + 3)
        ws.merge_cells(f"{col}11:{end_col}11")
        ws.merge_cells(f"{col}12:{end_col}12")
        ws[f"{col}11"] = "  " + label; ws[f"{col}11"].style = "kpi_label"
        ws[f"{col}12"] = formula;      ws[f"{col}12"].style = "kpi_val"
        ws[f"{col}12"].number_format = fmt

    # ---------- KPI row 2 (utilization — computed by VBA) ----------
    ws.row_dimensions[13].height = 18
    ws.row_dimensions[14].height = 34
    util_kpis = [
        ("B", "TOTAL RUN HRS",       "0.0"),
        ("F", "TOTAL DOWNTIME HRS",  "0.0"),
        ("J", "AVG UTIL (AVAILABILITY) %", "0.0"),
        ("N", "AVG UTIL (CALENDAR) %",     "0.0"),
    ]
    for col, label, fmt in util_kpis:
        end_col = get_column_letter(ord(col) - ord("A") + 3)
        ws.merge_cells(f"{col}13:{end_col}13")
        ws.merge_cells(f"{col}14:{end_col}14")
        ws[f"{col}13"] = "  " + label; ws[f"{col}13"].style = "kpi_label"
        ws[f"{col}14"] = 0
        ws[f"{col}14"].style = "kpi_val_pct" if "%" in label else "kpi_val"
        ws[f"{col}14"].number_format = fmt

    # ---------- ADD NEW MACHINE ----------
    ws.row_dimensions[16].height = 22
    ws.merge_cells("B16:P16")
    ws["B16"] = "  ADD  NEW  MACHINE"
    ws["B16"].style = "section"

    ws.row_dimensions[17].height = 20
    add_labels = ["Machine ID", "Machine Name", "Tag",
                  "Metric 1 Name", "Metric 2 Name", "Metric 3 Name"]
    add_cols = ["B", "D", "F", "H", "J", "L"]
    for lab, col in zip(add_labels, add_cols):
        end_col = get_column_letter(ord(col) - ord("A") + 2)
        ws.merge_cells(f"{col}17:{end_col}17")
        ws[f"{col}17"] = lab
        ws[f"{col}17"].style = "th_left"

    ws.row_dimensions[18].height = 26
    for col in add_cols:
        end_col = get_column_letter(ord(col) - ord("A") + 2)
        ws.merge_cells(f"{col}18:{end_col}18")
        ws[f"{col}18"].style = "input"
    ws.merge_cells("N18:P18")
    ws["N18"] = "▸  ADD MACHINE"; ws["N18"].style = "btn"

    # ---------- MACHINE picker + per-machine UTILIZATION table ----------
    ws.row_dimensions[20].height = 22
    ws.merge_cells("B20:P20")
    ws["B20"] = "  MACHINES  TO  COMPARE     +     PER-MACHINE  UTILIZATION"
    ws["B20"].style = "section"

    # Headers row 21
    ws.row_dimensions[21].height = 22
    ws["B21"] = "Show";          ws["B21"].style = "th"
    ws["C21"] = "Machine ID";    ws["C21"].style = "th"
    ws.merge_cells("D21:E21")
    ws["D21"] = "Machine Name";  ws["D21"].style = "th_left"
    ws.merge_cells("F21:G21")
    ws["F21"] = "Run Hrs";       ws["F21"].style = "th"
    ws.merge_cells("H21:I21")
    ws["H21"] = "Downtime Hrs";  ws["H21"].style = "th"
    ws.merge_cells("J21:L21")
    ws["J21"] = "Util Avail %";  ws["J21"].style = "th"
    ws.merge_cells("M21:P21")
    ws["M21"] = "Util Calendar %"; ws["M21"].style = "th"

    dv_show = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_show)

    for i in range(8):
        r = 22 + i
        ws.row_dimensions[r].height = 20
        ws[f"B{r}"].style = "pick_show"
        ws[f"C{r}"].style = "pick_id"
        ws.merge_cells(f"D{r}:E{r}"); ws[f"D{r}"].style = "value"
        ws.merge_cells(f"F{r}:G{r}"); ws[f"F{r}"].style = "util_val"
        ws.merge_cells(f"H{r}:I{r}"); ws[f"H{r}"].style = "util_val"
        ws.merge_cells(f"J{r}:L{r}"); ws[f"J{r}"].style = "util_val"
        ws.merge_cells(f"M{r}:P{r}"); ws[f"M{r}"].style = "util_val"
        dv_show.add(f"B{r}:B{r}")

    # Pre-populate first 4 slots with Line-A machines so the workbook shows data on open
    for i, (mid, mname) in enumerate([("M01", "Machine 01"), ("M02", "Machine 02"),
                                       ("M03", "Machine 03"), ("M04", "Machine 04")]):
        r = 22 + i
        ws[f"B{r}"] = "Yes"
        ws[f"C{r}"] = mid
        ws[f"D{r}"] = mname

    # Hidden last-tag tracker
    ws["AH1"] = "Line-A"

    # ---------- Comparison charts ----------
    ws.row_dimensions[31].height = 22
    ws.merge_cells("B31:P31")
    ws["B31"] = "  SIDE-BY-SIDE  METRIC  COMPARISON"
    ws["B31"].style = "section"

    # Staging area (hidden cols R:AF, header row 32, data rows 33+)
    ws["R32"] = "Date"; ws["R32"].style = "th"
    for i in range(1, 5):
        ws.cell(row=32, column=17 + i, value=f"M{i:02d}").style = "th"
    start = datetime(2026, 1, 1)
    for d in range(30):
        ws.cell(row=33 + d, column=18, value=start + timedelta(days=d)).number_format = "yyyy-mm-dd"
        for m in range(1, 5):
            ws.cell(row=33 + d, column=18 + m,
                    value=50 + m * 3 + (d % 7) - 3 + random.random() * 4)

    def make_chart(title, top_left, width=14):
        chart = LineChart()
        chart.title = title
        chart.style = 2
        chart.height = 8
        chart.width = width
        data = Reference(ws, min_col=19, max_col=22, min_row=32, max_row=62)
        cats = Reference(ws, min_col=18, min_row=33, max_row=62)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        for i, ser in enumerate(chart.series):
            ser.graphicalProperties.line.solidFill = [CHART_A, CHART_B, CHART_C, "D64545"][i]
            ser.graphicalProperties.line.width = 20000
            ser.smooth = True
        ws.add_chart(chart, top_left)

    make_chart("Metric 1 — comparison", "B32")
    make_chart("Metric 2 — comparison", "I32")
    make_chart("Metric 3 — comparison", "B49", width=28)

    # Footer
    ws.row_dimensions[65].height = 22
    ws.merge_cells("B65:P65")
    ws["B65"] = "  Ctrl+Shift+I = Import   |   Ctrl+Shift+R = Refresh   |   Ctrl+Shift+G = Generate Report"
    ws["B65"].style = "hdr_sub"

    # Hide staging cols
    for col in range(18, 35):
        ws.column_dimensions[get_column_letter(col)].hidden = True

    # Named ranges
    wb.defined_names["StartDate"]   = DefinedName("StartDate",   attr_text="Dashboard!$C$7")
    wb.defined_names["EndDate"]     = DefinedName("EndDate",     attr_text="Dashboard!$C$8")
    wb.defined_names["TagFilter"]   = DefinedName("TagFilter",   attr_text="Dashboard!$G$7")
    wb.defined_names["MetricFocus"] = DefinedName("MetricFocus", attr_text="Dashboard!$G$8")
    wb.defined_names["AddMID"]      = DefinedName("AddMID",      attr_text="Dashboard!$B$18")
    wb.defined_names["AddName"]     = DefinedName("AddName",     attr_text="Dashboard!$D$18")
    wb.defined_names["AddTag"]      = DefinedName("AddTag",      attr_text="Dashboard!$F$18")
    wb.defined_names["AddM1"]       = DefinedName("AddM1",       attr_text="Dashboard!$H$18")
    wb.defined_names["AddM2"]       = DefinedName("AddM2",       attr_text="Dashboard!$J$18")
    wb.defined_names["AddM3"]       = DefinedName("AddM3",       attr_text="Dashboard!$L$18")
    wb.defined_names["LastTag"]     = DefinedName("LastTag",     attr_text="Dashboard!$AH$1")
    # Utilization KPI targets
    wb.defined_names["KpiRunHrs"]   = DefinedName("KpiRunHrs",   attr_text="Dashboard!$B$14")
    wb.defined_names["KpiDownHrs"]  = DefinedName("KpiDownHrs",  attr_text="Dashboard!$F$14")
    wb.defined_names["KpiUtilAvail"]= DefinedName("KpiUtilAvail",attr_text="Dashboard!$J$14")
    wb.defined_names["KpiUtilCal"]  = DefinedName("KpiUtilCal",  attr_text="Dashboard!$N$14")


def build_reports(wb: Workbook) -> None:
    ws = wb.create_sheet("Reports")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = LIME
    ws.row_dimensions[1].height = 46
    ws.merge_cells("B1:I1")
    ws["B1"] = "REPORT  ARCHIVE"
    ws["B1"].style = "hdr_title"
    ws.merge_cells("B2:I2")
    ws["B2"] = "Every 'Generate Report' click appends a snapshot below and exports a PDF next to the workbook."
    ws["B2"].style = "hdr_sub"

    headers = ["Generated At", "Tag", "Start", "End", "Machines", "Points", "Report Sheet", "PDF File"]
    for i, h in enumerate(headers, start=2):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 26

    widths = {"A": 2, "B": 20, "C": 12, "D": 12, "E": 12, "F": 10, "G": 10, "H": 24, "I": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = STEEL
    ws.row_dimensions[1].height = 46
    ws.merge_cells("B1:J1")
    ws["B1"] = "SETUP  &  USAGE"
    ws["B1"].style = "hdr_title"

    steps = [
        ("1.", "ONE-TIME SETUP",  "Alt+F11 → Import File… → vba/MachineMonitor.bas. Save As Macro-Enabled Workbook (*.xlsm). Then Alt+F8 → SetupWorkbook → Run."),
        ("2.", "CONFIGURE",       "Config sheet holds up to 16 machines. Or add live from the Dashboard using the ADD MACHINE panel."),
        ("3.", "IMPORT DATA",     "Dashboard → IMPORT CSV. Columns: machine_id, tag, timestamp, metric1, metric2, metric3, running_hours(cum), downtime_hours(delta)."),
        ("4.", "FILTER",          "Set Start / End dates + Tag. Click REFRESH DASHBOARD."),
        ("5.", "PICK MACHINES",   "In 'MACHINES TO COMPARE', untick any machine to exclude, then REFRESH again."),
        ("6.", "UTILIZATION",     "Two formulas: Availability = Run/(Run+Down) × 100.  Calendar = Run/(24h × days in range) × 100."),
        ("7.", "REPORT",          "GENERATE REPORT creates a frozen snapshot sheet + a PDF next to the workbook."),
        ("8.", "SHORTCUTS",       "Ctrl+Shift+I Import   |   Ctrl+Shift+R Refresh   |   Ctrl+Shift+G Report."),
        ("9.", "LIVE UPDATES",    "Run ToggleLive to auto-refresh every 60 sec once you have a live feed."),
    ]
    for i, (n, title, body) in enumerate(steps, start=3):
        ws.cell(row=i, column=2, value=n).style = "kpi_label"
        ws.cell(row=i, column=3, value=title).style = "section"
        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=10)
        ws.cell(row=i, column=4, value=body).style = "label"

    widths = {"A": 2, "B": 4, "C": 22, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    make_styles(wb)
    build_config(wb)
    build_data(wb)
    build_reports(wb)
    build_readme_sheet(wb)
    build_dashboard(wb)
    wb.active = 0
    wb.save(OUT_XLSX)
    print(f"✔ Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
